from openpyxl import load_workbook
import numpy as np
import scipy.io as sio

def process_sheet(wb,sheet_name):
    current_sheet = wb[sheet_name]
    num_rows = current_sheet.max_row

    columns_of_interest = ['A','B','C','D','E','G','H','I','J','K','L','M','N']

    # first three rows are title etc.

    sample = np.zeros((num_rows-3,len(columns_of_interest)))



    for i in range(4,num_rows+1):

        current_row = []
        for j in range(len(columns_of_interest)):
            sample[i-4,j] = current_sheet[columns_of_interest[j]+str(i)].value




    return sample


def find_longest_sequence(sequences):

    max_length = 0

    for sequence in sequences:
        if(sequence.shape[0]>max_length):
            max_length = sequence.shape[0]

    return max_length




wb2 = load_workbook('20170103 - Strain 207 (Eco1-AID-TIR) - 1mM IAA - Position 0.xlsx')
sheet_names = wb2.get_sheet_names()


del sheet_names[0]

sequences = []

for sheet_name in sheet_names:
    sequences.append(process_sheet(wb2,sheet_name))
    print "File done"


max_sequence_length  = find_longest_sequence(sequences)


full_data = np.zeros((len(sheet_names),max_sequence_length,13))

for i in range(len(sheet_names)):
    full_data[i,:sequences[i].shape[0],:] = sequences[i]


sio.savemat('full_data_207_pos0.mat',{'data':full_data})


